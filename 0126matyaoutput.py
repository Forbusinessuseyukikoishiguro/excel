# pandas（データ分析ライブラリ）をpdとして読み込み
import pandas as pd

# openpyxlライブラリから、Excel操作に必要なWorkbookとload_workbookクラスを読み込み
from openpyxl import load_workbook, Workbook

# OSの機能（ファイルパス操作など）を利用するためのosモジュールを読み込み
import os

# Excel操作を行うためのメインクラスを定義
class ExcelExtractor:
    # クラスの初期化メソッド。複数の入力ファイルと1つの出力ファイルのパスを受け取る
    def __init__(self, input_files, output_file):
        """
        ExcelExtractorクラスの初期化
        
        Parameters:
        input_files (list): 入力ファイルパスのリスト
        output_file (str): 出力ファイルパス
        """
        # 入力ファイルのリストをインスタンス変数として保存
        self.input_files = input_files
        # 出力ファイルのパスをインスタンス変数として保存
        self.output_file = output_file
    
    # 抹茶商品を抽出して保存するメインメソッド
    def extract_and_save_matcha_items(self):
        """
        入力ファイルから抹茶商品を抽出し、新しいファイルに保存します
        """
        try:
            # 新しい出力用Excelワークブックを作成
            output_wb = Workbook()
            # アクティブシートを取得
            output_sheet = output_wb.active
            # シート名を設定
            output_sheet.title = "抹茶商品リスト"
            
            # 出力ファイルのヘッダー行を設定
            output_sheet['A1'] = "商品名"
            output_sheet['B1'] = "ファイル元"
            # データの書き込み開始行を設定（ヘッダーの次の行から）
            current_row = 2
            
            # 入力ファイルを1つずつ処理
            for input_file in self.input_files:
                # 入力ファイルを読み込み
                input_wb = load_workbook(input_file)
                # アクティブシートを取得
                input_sheet = input_wb.active
                # ファイル名のみを取得（パスを除く）
                file_name = os.path.basename(input_file)
                
                # シート内の全ての行をループ処理
                for row in input_sheet.iter_rows():
                    # 行内の各セルをループ処理
                    for cell in row:
                        # セルが空でなく、文字列型で、「抹茶」を含む場合
                        if cell.value and isinstance(cell.value, str) and "抹茶" in cell.value:
                            # 抽出した商品名を出力ファイルに書き込み
                            output_sheet[f'A{current_row}'] = cell.value
                            # 元ファイル名を出力ファイルに書き込み
                            output_sheet[f'B{current_row}'] = file_name
                            # 抽出状況を画面に表示
                            print(f"抽出: {cell.value} (from {file_name})")
                            # 次の行番号に進む
                            current_row += 1
            
            # 出力ファイルを保存
            output_wb.save(self.output_file)
            # 処理結果を画面に表示（current_row - 2 でヘッダー行を除いた件数を計算）
            print(f"\n合計 {current_row - 2} 件の抹茶商品を抽出し、{self.output_file} に保存しました。")
            
        # エラーが発生した場合の処理
        except Exception as e:
            # エラーメッセージを画面に表示
            print(f"エラーが発生しました: {str(e)}")

# このファイルが直接実行された場合のみ実行される部分
if __name__ == "__main__":
    # 基本となるフォルダパスを設定
    base_path = r"C:\Users\yukik\Desktop\excel"
    # 入力ファイル1のパスを設定
    input_file1 = os.path.join(base_path, "ex1.xlsx")
    # 入力ファイル2のパスを設定
    input_file2 = os.path.join(base_path, "ex2.xlsx")
    # 出力ファイルのパスを設定
    output_file = os.path.join(base_path, "exoutput.xlsx")
    
    # ExcelExtractorクラスのインスタンスを作成（入力ファイルリストと出力ファイルを指定）
    extractor = ExcelExtractor([input_file1, input_file2], output_file)
    # 抹茶商品の抽出と保存を実行
    extractor.extract_and_save_matcha_items()
#2025/1/26解説丁寧版　抹茶と記載があるものを抽出し、exoutput.xlsxに保存しましたOK インプットはex1とex2の２ファイル