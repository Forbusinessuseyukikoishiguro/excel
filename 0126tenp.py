# pandas: データ分析や表形式データを扱うための強力なライブラリをインポート。Excelファイルの読み書きに便利
import pandas as pd

# openpyxl: Excel操作専用のライブラリから必要なクラスをインポート
# load_workbook: 既存のExcelファイルを読み込むためのクラス
# Workbook: 新しいExcelファイルを作成するためのクラス
from openpyxl import load_workbook, Workbook

# os: ファイルやディレクトリの操作を行うための標準ライブラリ
import os

# datetime: 日付と時刻を扱うためのライブラリ。ファイル名に時刻を付けるなどの用途に使用
from datetime import datetime

# Excelファイルの処理を行うメインクラス。複数の機能をまとめて管理
class ExcelProcessor:
    def __init__(self):
        """クラスの初期化メソッド。インスタンス変数を初期化"""
        # テンプレートデータを保存する変数を初期化
        self.template_data = None
        # 挿入データを保存する変数を初期化
        self.insert_data = None
        
    def read_template(self, template_path):
        """
        テンプレートファイルを読み込むメソッド
        template_path: テンプレートファイルのパス
        """
        try:
            # テンプレートファイルをExcelワークブックとして読み込み
            workbook = load_workbook(template_path)
            # 読み込んだデータをインスタンス変数に保存
            self.template_data = workbook
            # 成功メッセージを表示
            print(f"テンプレートを読み込みました: {os.path.basename(template_path)}")
            return True
        except Exception as e:
            # エラーが発生した場合のメッセージを表示
            print(f"テンプレート読み込み中にエラーが発生: {str(e)}")
            return False

    def read_insert_data(self, insert_data_path):
        """
        挿入するデータファイルを読み込むメソッド
        insert_data_path: 挿入データファイルのパス
        """
        try:
            # データファイルをExcelワークブックとして読み込み
            workbook = load_workbook(insert_data_path)
            # 読み込んだデータをインスタンス変数に保存
            self.insert_data = workbook
            # 成功メッセージを表示
            print(f"挿入データを読み込みました: {os.path.basename(insert_data_path)}")
            return True
        except Exception as e:
            # エラーが発生した場合のメッセージを表示
            print(f"データ読み込み中にエラーが発生: {str(e)}")
            return False

    def create_output_from_template(self, output_path):
        """
        テンプレートを基に出力ファイルを作成するメソッド
        output_path: 出力ファイルのパス
        """
        try:
            # テンプレートデータが読み込まれているか確認
            if self.template_data is None:
                raise Exception("テンプレートが読み込まれていません")
            
            # テンプレートを新しいファイルとして保存
            self.template_data.save(output_path)
            # 成功メッセージを表示
            print(f"テンプレートを出力しました: {os.path.basename(output_path)}")
            return True
        except Exception as e:
            # エラーが発生した場合のメッセージを表示
            print(f"テンプレート出力中にエラーが発生: {str(e)}")
            return False

    def insert_data_to_output(self, output_path):
        """
        出力ファイルにデータを挿入するメソッド
        output_path: データを挿入する出力ファイルのパス
        """
        try:
            # 挿入データが読み込まれているか確認
            if self.insert_data is None:
                raise Exception("挿入データが読み込まれていません")
            
            # 出力ファイルを読み込む
            output_wb = load_workbook(output_path)
            # アクティブ（現在選択されている）シートを取得
            output_sheet = output_wb.active
            
            # 挿入データのアクティブシートを取得
            insert_sheet = self.insert_data.active
            
            # データを挿入（A1セルから開始）
            for row_idx, row in enumerate(insert_sheet.iter_rows(), 1):
                for col_idx, cell in enumerate(row, 1):
                    # セルの値がNoneでない場合のみ書き込み
                    if cell.value is not None:
                        output_sheet.cell(row=row_idx, column=col_idx, value=cell.value)
            
            # 変更を保存
            output_wb.save(output_path)
            # 成功メッセージを表示
            print(f"データを挿入して保存しました: {os.path.basename(output_path)}")
            return True
        except Exception as e:
            # エラーが発生した場合のメッセージを表示
            print(f"データ挿入中にエラーが発生: {str(e)}")
            return False

# メインの実行部分。プログラムが直接実行された時のみ実行される
def main():
    # 基本となるフォルダのパスを設定
    base_path = r"C:\Users\yukik\Desktop\excel\all"
    # 各ファイルのパスを設定
    template_path = os.path.join(base_path, "ex1_template.xlsx")
    insert_data_path = os.path.join(base_path, "ex2_insertdata.xlsx")
    output_path = os.path.join(base_path, "ex3_all_output.xlsx")
    
    # ExcelProcessorクラスのインスタンスを作成
    processor = ExcelProcessor()
    
    # 処理開始メッセージを表示
    print("\n=== 処理を開始します ===")
    
    # 1. テンプレートの読み込み
    print("\nステップ1: テンプレートの読み込み")
    if not processor.read_template(template_path):
        return
    
    # 2. 挿入データの読み込み
    print("\nステップ2: 挿入データの読み込み")
    if not processor.read_insert_data(insert_data_path):
        return
    
    # 3. テンプレートを基に出力ファイルを作成
    print("\nステップ3: 出力ファイルの作成")
    if not processor.create_output_from_template(output_path):
        return
    
    # 4. 出力ファイルにデータを挿入
    print("\nステップ4: データの挿入")
    if not processor.insert_data_to_output(output_path):
        return
    
    # 処理完了メッセージを表示
    print("\n=== すべての処理が完了しました ===")

# プログラムのエントリーポイント。このファイルが直接実行された場合のみmain()を実行
if __name__ == "__main__":
    main()
    
# このプログラムは、指定されたテンプレートファイルと挿入データファイルを読み込み、アウトプットファイルにデータを挿入する処理を行います。
#202/01/26ExcelOK