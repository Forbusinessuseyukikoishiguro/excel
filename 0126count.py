# pandasライブラリをpdという略称でインポート。データ分析や表形式データの処理に使用
import pandas as pd

# openpyxlライブラリからload_workbookクラスをインポート。Excelファイルの読み書きに使用
from openpyxl import load_workbook

# osモジュールをインポート。ファイルやディレクトリの操作に使用
import os

# ExcelProcessorクラスを定義。Excelファイルの処理を担当
class ExcelProcessor:
    # クラスの初期化メソッド。ファイルパスを受け取り、インスタンス変数として保存
    def __init__(self, file_path):
        # インスタンス変数file_pathにファイルパスを保存
        self.file_path = file_path
    
    # キーワードを検索し、隣のセルにメッセージを追加するメソッド
    def add_sale_message(self, keyword, message="お買い得！"):
        """
        指定したキーワードを含むセルの隣のセルにメッセージを追加します
        
        Parameters:
        keyword (str): 検索するキーワード
        message (str): 追加するメッセージ（デフォルト: お買い得！）
        """
        try:
            # Excelファイルを読み込み
            workbook = load_workbook(self.file_path)
            # アクティブなシートを取得
            sheet = workbook.active
            # 更新したセルの数を記録する変数を初期化
            update_count = 0
            
            # シート内の全ての行に対してループ
            for row in sheet.iter_rows():
                # 行内の各セルに対してループ
                for cell in row:
                    # セルが空でなく、文字列型で、指定したキーワードを含む場合
                    if cell.value and isinstance(cell.value, str) and keyword in cell.value:
                        # 現在のセルの右隣のセルを取得（列番号+1）
                        next_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                        # 右隣のセルにメッセージを書き込み
                        next_cell.value = message
                        # 更新カウントを増やす
                        update_count += 1
                        # 更新内容を表示
                        print(f"{cell.coordinate}の隣のセル{next_cell.coordinate}に'{message}'を追加しました")
            
            # 変更をファイルに保存
            workbook.save(self.file_path)
            # 更新完了メッセージを表示
            print(f"\n合計{update_count}箇所を更新し、保存しました")
            
        # エラーが発生した場合の処理
        except Exception as e:
            # エラーメッセージを表示
            print(f"エラーが発生しました: {str(e)}")

# このファイルが直接実行された場合にのみ実行される部分
if __name__ == "__main__":
    # 処理対象のExcelファイルのパスを指定
    input_file1 = r"C:\Users\yukik\Desktop\excel\ex1.xlsx"
    
    # ExcelProcessorクラスのインスタンスを作成
    processor = ExcelProcessor(input_file1)
    # 「抹茶」というキーワードで検索し、隣のセルに「お買い得！」を追加
    processor.add_sale_message("抹茶")