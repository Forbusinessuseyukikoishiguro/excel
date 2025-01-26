import pandas as pd
from openpyxl import load_workbook, Workbook
import os
import shutil
from datetime import datetime

class ExcelBatchTester:
    def __init__(self, base_path):
        """
        バッチ処理テスト用クラスの初期化
        
        Parameters:
        base_path (str): テスト用ファイルの基本パス
        """
        self.base_path = base_path
        self.test_dir = os.path.join(base_path, "test_" + datetime.now().strftime("%Y%m%d_%H%M%S"))
        self.test_results = []
        
    def setup_test_environment(self):
        """テスト環境のセットアップ"""
        try:
            # テストディレクトリの作成
            os.makedirs(self.test_dir, exist_ok=True)
            
            # テスト用ファイルのコピー
            source_files = ["ex1.xlsx", "ex2.xlsx"]
            for file in source_files:
                src = os.path.join(self.base_path, file)
                dst = os.path.join(self.test_dir, file)
                shutil.copy2(src, dst)
            
            self.log_result("環境セットアップ", "テスト環境を作成しました", True)
        except Exception as e:
            self.log_result("環境セットアップ", f"エラー: {str(e)}", False)
    
    def test_write_to_excel(self):
        """Excelファイルへの書き込みテスト"""
        try:
            test_file = os.path.join(self.test_dir, "ex1.xlsx")
            wb = load_workbook(test_file)
            sheet = wb.active
            sheet['D1'] = 'テストデータ'
            wb.save(test_file)
            self.log_result("Excel書き込み", "データを書き込みました", True)
        except Exception as e:
            self.log_result("Excel書き込み", f"エラー: {str(e)}", False)
    
    def test_read_from_excel(self):
        """Excelファイルからの読み込みテスト"""
        try:
            test_file = os.path.join(self.test_dir, "ex1.xlsx")
            wb = load_workbook(test_file)
            sheet = wb.active
            value = sheet['D1'].value
            if value == 'テストデータ':
                self.log_result("Excel読み込み", "データを正常に読み込みました", True)
            else:
                self.log_result("Excel読み込み", "データが一致しません", False)
        except Exception as e:
            self.log_result("Excel読み込み", f"エラー: {str(e)}", False)
    
    def test_search_keyword(self):
        """キーワード検索テスト"""
        try:
            test_file = os.path.join(self.test_dir, "ex1.xlsx")
            wb = load_workbook(test_file)
            sheet = wb.active
            found = False
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "抹茶" in cell.value:
                        found = True
                        break
            if found:
                self.log_result("キーワード検索", "「抹茶」を含むセルを見つけました", True)
            else:
                self.log_result("キーワード検索", "「抹茶」を含むセルが見つかりません", False)
        except Exception as e:
            self.log_result("キーワード検索", f"エラー: {str(e)}", False)
    
    def log_result(self, test_name, message, success):
        """テスト結果のログを記録"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        result = "成功" if success else "失敗"
        self.test_results.append({
            "タイムスタンプ": timestamp,
            "テスト名": test_name,
            "結果": result,
            "メッセージ": message
        })
    
    def save_test_results(self):
        """テスト結果をExcelファイルに保存"""
        try:
            # テスト結果をDataFrameに変換
            df = pd.DataFrame(self.test_results)
            
            # 結果をExcelファイルとして保存
            result_file = os.path.join(self.test_dir, "test_results.xlsx")
            df.to_excel(result_file, index=False)
            print(f"\nテスト結果を保存しました: {result_file}")
        except Exception as e:
            print(f"テスト結果の保存中にエラーが発生しました: {str(e)}")
    
    def run_all_tests(self):
        """すべてのテストを実行"""
        print("バッチ処理テストを開始します...")
        self.setup_test_environment()
        self.test_write_to_excel()
        self.test_read_from_excel()
        self.test_search_keyword()
        self.save_test_results()
        print("\nテスト完了")
        
        # テスト結果の表示
        print("\nテスト結果サマリー:")
        for result in self.test_results:
            print(f"{result['テスト名']}: {result['結果']} - {result['メッセージ']}")

if __name__ == "__main__":
    base_path = r"C:\Users\yukik\Desktop\excel"
    tester = ExcelBatchTester(base_path)
    tester.run_all_tests()
#2025/01/26 バッチ処理テストを開始します...