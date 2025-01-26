# pandasライブラリをインポート：データフレーム操作のための主要なデータ分析ツール
import pandas as pd

# openpyxlライブラリから必要なクラスをインポート：Excel操作の基本機能を提供
from openpyxl import load_workbook, Workbook

# osモジュールをインポート：ファイルシステム操作（パス結合、ディレクトリ作成など）に使用
import os

# datetimeモジュールをインポート：タイムスタンプ生成やファイル名への時刻付与に使用
from datetime import datetime

# Excel行分割処理を行うメインクラスを定義
class ExcelRowSplitter:
    def __init__(self, input_files, output_dir):
        """
        行分割処理クラスの初期化メソッド
        
        Parameters:
        input_files (list): 入力ファイルパスのリスト
        output_dir (str): 出力ディレクトリのパス
        """
        # 入力ファイルのパスリストをインスタンス変数として保存
        self.input_files = input_files
        # 出力ディレクトリのパスをインスタンス変数として保存
        self.output_dir = output_dir
        # 抽出したデータを保存するリストを初期化
        self.all_data = []

    def extract_data_from_excel(self):
        """Excelファイルからデータを抽出するメソッド"""
        try:
            # 入力ファイルを1つずつ処理
            for file_path in self.input_files:
                # Excelファイルを読み込み
                wb = load_workbook(file_path)
                # アクティブシートを取得
                sheet = wb.active
                
                # シート内の各行を処理
                for row in sheet.iter_rows():
                    # 行データを格納するリストを初期化
                    row_data = []
                    # 行内の各セルを処理
                    for cell in row:
                        # セルの値がNoneの場合は空文字に変換して追加
                        row_data.append(cell.value if cell.value is not None else "")
                    # 行が完全に空でない場合のみデータとして追加
                    if any(row_data):
                        self.all_data.append(row_data)
                            
            # 抽出したデータの行数を表示
            print(f"合計 {len(self.all_data)} 行のデータを抽出しました")
            
        # エラー発生時の処理
        except Exception as e:
            print(f"データ抽出中にエラーが発生しました: {str(e)}")

    def split_and_save_data(self, rows_per_file=3):
        """データを指定行数で分割して保存するメソッド"""
        try:
            # データを指定行数（デフォルト3行）ごとに分割して処理
            for i in range(0, len(self.all_data), rows_per_file):
                # 現在の位置から指定行数分のデータを取得
                chunk = self.all_data[i:i + rows_per_file]
                
                # 新しいExcelワークブックを作成
                wb = Workbook()
                # アクティブシートを取得
                sheet = wb.active
                
                # 分割したデータを新しいシートに書き込み
                for row_idx, row_data in enumerate(chunk, 1):
                    # 各行のデータを列ごとに書き込み
                    for col_idx, value in enumerate(row_data, 1):
                        sheet.cell(row=row_idx, column=col_idx, value=value)
                
                # ファイル番号を計算（1から開始）
                file_number = (i // rows_per_file) + 1
                # 出力ファイル名を生成（3桁の連番付き）
                output_file = os.path.join(self.output_dir, f"split_data_{file_number:03d}.xlsx")
                # ファイルを保存
                wb.save(output_file)
                # 保存完了メッセージを表示
                print(f"保存完了: {output_file}")
                
        # エラー発生時の処理
        except Exception as e:
            print(f"データ分割・保存中にエラーが発生しました: {str(e)}")

    def process_files(self):
        """一連の処理を実行するメインメソッド"""
        # 処理開始メッセージを表示
        print("処理を開始します...")
        # 出力ディレクトリを作成（既に存在する場合は何もしない）
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Excelファイルからデータを抽出
        self.extract_data_from_excel()
        
        # 抽出したデータを分割して保存
        self.split_and_save_data()
        
        # 処理完了メッセージを表示
        print("すべての処理が完了しました")

# このファイルが直接実行された場合にのみ実行される部分
if __name__ == "__main__":
    # 基本となるフォルダパスを設定
    base_path = r"C:\Users\yukik\Desktop\excel"
    
    # 入力ファイルのパスを設定
    input_file1 = os.path.join(base_path, "ex1.xlsx")
    input_file2 = os.path.join(base_path, "ex2.xlsx")
    
    # 現在時刻を含む出力ディレクトリ名を生成
    output_dir = os.path.join(base_path, "split_files_" + datetime.now().strftime("%Y%m%d_%H%M%S"))
    
    # ExcelRowSplitterのインスタンスを作成し、処理を実行
    splitter = ExcelRowSplitter([input_file1, input_file2], output_dir)
    splitter.process_files()
    
    #2025/01/26 3行ごとに分割して保存処理＿まとめるテスト_インプット自動保存を統合して実行して入力（まとめてExcel２ファイル入力処理OK！）コメント付与